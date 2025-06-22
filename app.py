from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import requests
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import io

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'una_clave_secreta_muy_segura_aqui') # ¡IMPORTANTE: Cambia esto en Render!

# --- Función de procesamiento de fecha/hora ---
def convertir_fecha_hora(fecha_hora_str):
    meses = {
        "January": "01", "February": "02", "March": "03", "April": "04",
        "May": "05", "June": "06", "July": "07", "August": "08",
        "September": "09", "October": "10", "November": "11", "December": "12"
    }
    match = re.match(r"(\d{1,2}) de ([a-zA-Z]+) de (\d{4}) en horas: (\d{2}:\d{2}:\d{2})", fecha_hora_str)
    if match:
        dia, mes, anio, hora = match.groups()
        mes_num = meses.get(mes, "00")
        return f"{dia.zfill(2)}/{mes_num}/{anio} {hora}"
    return fecha_hora_str

@app.route('/', methods=['GET', 'POST'])
def iniciar_formulario():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']

        login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"
        data_url = "http://sigof.distriluz.com.pe/plus/ComlecOrdenlecturas/ajax_mostar_mapa_selfie"

        credentials = {
            "data[Usuario][usuario]": usuario,
            "data[Usuario][pass]": password
        }

        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": login_url,
        }

        with requests.Session() as session:
            login_response = session.post(login_url, data=credentials, headers=headers)
            if "Usuario o contraseña incorrecto" in login_response.text:
                flash("🧠 Humano, las credenciales son incorrectas.", 'error')
                return redirect(url_for('iniciar_formulario'))
            
            data_response = session.get(data_url, headers=headers)

        data = data_response.text
        data_cleaned = data.replace("\\/", "/")
        data_cleaned = re.sub(r"<\/?\w+.*?>", "", data_cleaned)
        data_cleaned = re.sub(r"\s+", " ", data_cleaned).strip()
        blocks = re.split(r"Ver detalle", data_cleaned)

        results = {}
        for block in blocks:
            fecha = re.search(r"Fecha Selfie:\s*(\d{1,2} de [a-zA-Z]+ de \d{4} en horas: \d{2}:\d{2}:\d{2})", block)
            lecturista = re.search(r"Lecturista:\s*([\w\sÁÉÍÓÚáéíóúÑñ]+)", block)
            url = re.search(r"url\":\"(https[^\"]+)", block)

            if fecha and lecturista and url:
                fecha_hora_formateada = convertir_fecha_hora(fecha.group(1).strip())
                fecha_selfie, _ = fecha_hora_formateada.split(" ")
                lecturista_nombre = lecturista.group(1).strip()
                url_imagen = url.group(1).strip()

                key = (lecturista_nombre, fecha_selfie)
                if key not in results:
                    results[key] = {"URLs Imagen": []}
                results[key]["URLs Imagen"].append(url_imagen)

        if results:
            max_urls = max(len(item["URLs Imagen"]) for item in results.values())
            url_columns = [f"Url_foto {i+1}" for i in range(max_urls)]
            columns = ["Fecha Selfie", "Lecturista"] + url_columns
            vista_columns = [f"Vista Url_foto {i+1}" for i in range(max_urls)]

            data_list = []
            for (lecturista, fecha_selfie), info in results.items():
                row = [fecha_selfie, lecturista] + info["URLs Imagen"] + [""] * (max_urls - len(info["URLs Imagen"]))
                data_list.append(row)

            df = pd.DataFrame(data_list, columns=columns)

            # Crear el archivo Excel en memoria
            wb = Workbook()
            ws = wb.active
            ws.title = "LmcSelfiesLectura"
            ws.append(columns + vista_columns)

            for i, row in enumerate(df.itertuples(index=False), start=2):
                row_data = list(row)
                ws.append(row_data + [""] * max_urls)

                for j in range(max_urls):
                    col_index = 3 + j
                    url_cell = f"{get_column_letter(col_index)}{i}"
                    vista_col_index = len(columns) + j + 1
                    formula_cell = f"{get_column_letter(vista_col_index)}{i}"

                    image_height_px = 200
                    image_width_px = 140
                    ws[formula_cell] = f'=IMAGE({url_cell},4,{image_height_px},{image_width_px})'

                    vista_col_letter = get_column_letter(vista_col_index)
                    ws.column_dimensions[vista_col_letter].width = round(image_width_px / 7, 1)

                ws.row_dimensions[i].height = 151
            
            # Guardar el libro en un buffer en memoria
            excel_file_buffer = io.BytesIO()
            wb.save(excel_file_buffer)
            excel_file_buffer.seek(0) # Volver al inicio del buffer

            filename = "Lmc_ReporteSelfie.xlsx"
            
            # Devolver el archivo Excel para descarga
            flash("✅ Humano, tu archivo de reporte de selfies está listo para descargar.", 'success')
            return send_file(
                excel_file_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        else:
            flash("⚠️ Humano, tu usuario o contraseña es incorrecta / no se encontró datos para exportar.", 'warning')
            return redirect(url_for('iniciar_formulario'))

    # Si es una solicitud GET, simplemente muestra el formulario
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
