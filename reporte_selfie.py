def generar_excel_selfies():
    # -------------------------------------
    # INSTALAR Y CONFIGURAR GOOGLE DRIVE
    # -------------------------------------
    from google.colab import auth
    auth.authenticate_user()

    from pydrive.auth import GoogleAuth
    from pydrive.drive import GoogleDrive
    from oauth2client.client import GoogleCredentials

    gauth = GoogleAuth()
    gauth.credentials = GoogleCredentials.get_application_default()
    drive = GoogleDrive(gauth)

    # -------------------------------------
    # FORMULARIO PARA INGRESAR CREDENCIALES
    # -------------------------------------
    from IPython.display import display, HTML, clear_output
    import ipywidgets as widgets
    import requests
    import re
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    caja_tamano = '250px'

    usuario_input = widgets.Text(
        value='',
        placeholder='Ingrese su usuario',
        description='👤 Usuario:',
        style={'description_width': 'initial'},
        layout=widgets.Layout(width=caja_tamano)
    )

    password_input = widgets.Password(
        value='',
        placeholder='Ingrese su contraseña',
        description='🔑 Clave:',
        style={'description_width': 'initial'},
        layout=widgets.Layout(width=caja_tamano)
    )

    boton = widgets.Button(
        description='🔓 Humano inicia sesión',
        button_style='primary',
        layout=widgets.Layout(width='200px')
    )

    mensaje_error = widgets.HTML(
        value='',
        layout=widgets.Layout(visibility='hidden', margin='10px 0 0 0')
    )

    formulario = widgets.VBox([
        widgets.HBox([usuario_input], layout=widgets.Layout(justify_content='flex-end', width=caja_tamano)),
        widgets.HBox([password_input], layout=widgets.Layout(justify_content='flex-end', width=caja_tamano)),
        widgets.HBox([boton], layout=widgets.Layout(justify_content='center', width='200px')),
        mensaje_error
    ])

    # TÍTULO CON FONDO CELESTE ADAPTADO AL CONTENIDO
    display(HTML("""
    <div style="
        display: inline-block;
        background-color: #007bff;
        padding: 5px 10px;
        border-radius: 8px;
        font-weight: bold;
        font-size: 14px;
        color: white; /* Color blanco */
        box-shadow: 1px 1px 4px rgba(0,0,0,0.1);
    ">
    📋 HUMANO INGRESA TU USUARIO Y CONTRASEÑA DE SIGOF WEB
    </div>
    """))

    display(formulario)

    def on_login_clicked(b):
        login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"
        data_url = "http://sigof.distriluz.com.pe/plus/ComlecOrdenlecturas/ajax_mostar_mapa_selfie"

        credentials = {
            "data[Usuario][usuario]": usuario_input.value,
            "data[Usuario][pass]": password_input.value
        }

        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": login_url,
        }

        with requests.Session() as session:
            login_response = session.post(login_url, data=credentials, headers=headers)
            if "Usuario o contraseña incorrecto" in login_response.text:
                mensaje_error.value = "<span style='color: red;'>🧠 Humano, las credenciales son incorrectas.</span>"
                mensaje_error.layout.visibility = 'visible'
                return
            else:
                mensaje_error.value = ""
                mensaje_error.layout.visibility = 'hidden'

            data_response = session.get(data_url, headers=headers)

        data = data_response.text
        data_cleaned = data.replace("\\/", "/")
        data_cleaned = re.sub(r"<\/?\w+.*?>", "", data_cleaned)
        data_cleaned = re.sub(r"\s+", " ", data_cleaned).strip()
        blocks = re.split(r"Ver detalle", data_cleaned)

        def convertir_fecha_hora(fecha_hora_str):
            meses = {
                "January": "01", "February": "02", "March": "03", "April": "04",
                "May": "05", "June": "06", "July": "07", "August": "08",
                "September": "09", "October": "10", "November": "11", "December": "12"
            }
            match = re.match(r"(\d{1,2}) de ([a-zA-Z]+) de (\d{4}) en horas: (\d{2}:\d{2}:\d{2})", fecha_hora_str)
            if match:
                dia, mes, anio, hora = match.groups()
                mes_num = meses.get(mes, "00")
                return f"{dia.zfill(2)}/{mes_num}/{anio} {hora}"
            return fecha_hora_str

        results = {}
        for block in blocks:
            fecha = re.search(r"Fecha Selfie:\s*(\d{1,2} de [a-zA-Z]+ de \d{4} en horas: \d{2}:\d{2}:\d{2})", block)
            lecturista = re.search(r"Lecturista:\s*([\w\sÁÉÍÓÚáéíóúÑñ]+)", block)
            url = re.search(r"url\":\"(https[^\"]+)", block)

            if fecha and lecturista and url:
                fecha_hora_formateada = convertir_fecha_hora(fecha.group(1).strip())
                fecha_selfie, _ = fecha_hora_formateada.split(" ")
                lecturista_nombre = lecturista.group(1).strip()
                url_imagen = url.group(1).strip()

                key = (lecturista_nombre, fecha_selfie)
                if key not in results:
                    results[key] = {"URLs Imagen": []}
                results[key]["URLs Imagen"].append(url_imagen)

        if results:
            max_urls = max(len(item["URLs Imagen"]) for item in results.values())
            url_columns = [f"Url_foto {i+1}" for i in range(max_urls)]
            columns = ["Fecha Selfie", "Lecturista"] + url_columns
            vista_columns = [f"Vista Url_foto {i+1}" for i in range(max_urls)]

            data = []
            for (lecturista, fecha_selfie), info in results.items():
                row = [fecha_selfie, lecturista] + info["URLs Imagen"] + [""] * (max_urls - len(info["URLs Imagen"]))
                data.append(row)

            df = pd.DataFrame(data, columns=columns)

            wb = Workbook()
            ws = wb.active
            ws.title = "LmcSelfiesLectura"
            ws.append(columns + vista_columns)

            for i, row in enumerate(df.itertuples(index=False), start=2):
                row_data = list(row)
                ws.append(row_data + [""] * max_urls)

                for j in range(max_urls):
                    col_index = 3 + j
                    url_cell = f"{get_column_letter(col_index)}{i}"
                    vista_col_index = len(columns) + j + 1
                    formula_cell = f"{get_column_letter(vista_col_index)}{i}"

                    image_height_px = 200
                    image_width_px = 140
                    ws[formula_cell] = f'=IMAGE({url_cell})'

                    vista_col_letter = get_column_letter(vista_col_index)
                    ws.column_dimensions[vista_col_letter].width = round(image_width_px / 7, 1)

                ws.row_dimensions[i].height = 151

            filename = "Lmc_ReporteSelfie.xlsx"
            wb.save(filename)

            folder_name = "LMC REPORTE SELFIES"
            folder_id = None
            file_list = drive.ListFile({
                'q': f"title='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            }).GetList()

            if file_list:
                folder_id = file_list[0]['id']
            else:
                folder_metadata = {
                    'title': folder_name,
                    'mimeType': 'application/vnd.google-apps.folder'
                }
                folder = drive.CreateFile(folder_metadata)
                folder.Upload()
                folder_id = folder['id']

            gfile = drive.CreateFile({
                'title': filename,
                'parents': [{'id': folder_id}]
            })
            gfile.SetContentFile(filename)
            gfile.Upload()

            print(f"✅ Humano tu archivo se ha subido a la carpeta '{folder_name}' con éxito.")
            folder_link = f"https://drive.google.com/drive/folders/{folder_id}"
            display(HTML(f'<a href="{folder_link}" target="_blank"><button>📁  Abrir Lmc Reporte Selfies</button></a>'))

        else:
            print("⚠️ Humano tu usuario o contraseña es incorrecta / no se encontró datos para exportar.")

    boton.on_click(on_login_clicked)
